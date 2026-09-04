"""Render the workbook as a single-page HTML view.

The workbook is the deliverable; this is the readable version of it — one page,
scannable, no tab-switching. It reads the .xlsx a run just wrote and renders it,
recomputing nothing. If a number here disagrees with the spreadsheet, this file
is the one that is wrong.

NOT for publishing. The page contains named contacts and email addresses from
your CRM. It is a local file to open, hand over, or attach — the rule about
contact data staying inside your own environment applies to it exactly as it
does to the workbook.

The page title comes from the account row, never from a literal. An
organisation's name in the code is the fork this repository bans, however
cosmetic.
"""

from __future__ import annotations

import html
import os
import re

from openpyxl import load_workbook

from .coverage import NOT_ASSESSED
from .stakeholders import ICP_NOT_ASSESSED

# Carried over verbatim so the artifact looks the same week to week.
CSS = """
:root{color-scheme:light dark} body{font:14px/1.5 -apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;padding:22px;max-width:1180px;margin:auto;color:#1a1a1a;background:#fff}
@media(prefers-color-scheme:dark){body{color:#e7e7e7;background:#141414}}
h1{font-size:20px;margin:0 0 4px} .meta{color:#777;font-size:13px;margin-bottom:22px}
section{margin:0 0 32px} h2{font-size:16px;margin:0 0 2px} .cnt{display:inline-block;background:#2F5B7C;color:#fff;border-radius:10px;padding:1px 9px;font-size:12px;margin-left:6px}
.sub{color:#777;font-size:12px;margin:2px 0 9px} .foot{color:#999;font-size:11.5px;font-style:italic;margin-top:6px}
table{border-collapse:collapse;width:100%;font-size:12.5px} th{background:#2F5B7C;color:#fff;text-align:left;padding:6px 8px}
td{padding:4px 8px;border-bottom:1px solid #e4e4e4;vertical-align:top;max-width:240px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
@media(prefers-color-scheme:dark){td{border-bottom:1px solid #2a2a2a}}
.no{color:#c0392b;font-weight:700} .fl{color:#8a6d00;background:#f7e8a0;border-radius:6px;padding:0 6px;font-size:11px}
td.ok{background:#1e8e3e;color:#fff;font-weight:700;text-align:center} td.rej{color:#b06000;font-size:11px}
td.na{color:#8a8a8a;font-style:italic;font-size:11px}
a{color:inherit;text-decoration:none;border-bottom:1px dotted #999}
"""

# Sheet name -> section heading. The blurb under each heading is NOT written
# here: it is the sheet's own footnote rows, lifted to the top. One source of
# words, so the page cannot drift from the workbook the way a hardcoded second
# copy did.
SECTIONS = [
    ("3 - Company coverage", "Company coverage"),
    ("4 - Stakeholder list", "Stakeholder list"),
    # "Not in CRM", not "Not in HubSpot or Salesforce": the heading must not
    # name a system this run never called, and the sheet it renames is already
    # correctly titled. A vendor name in customer-facing output is the same
    # defect as a vendor field name in a query.
    ("2 - Missing from CRM", "Not in CRM"),
    ("1 - Met this week", "Met this week"),
]


def cell(col: str, value: str) -> tuple[str, str]:
    """(css class, inner html) for one cell, mirroring the workbook's emphasis."""
    v = html.escape(value)
    if not v:
        return "", ""
    # A test that never ran is neither a pass nor a rejection. Styling it as
    # either — green, or the orange every other non-"yes" verdict gets — would
    # put back in colour the conflation the value itself removes. Matched
    # against the constants rather than a copy of the words, so the page cannot
    # drift from the workbook.
    if value in (NOT_ASSESSED, ICP_NOT_ASSESSED):
        return "na", v
    if col == "Meets profile?":
        return ("ok", v) if v == "yes" else ("rej", v)
    if v in ("NO", "GAP"):
        return "", f'<span class="no">{v}</span>'
    if col == "Recent contact?":
        return ("", v) if v.startswith("yes") else ("", f'<span class="fl">{v}</span>')
    if col == "Flag":
        return "", f'<span class="fl">{v}</span>'
    if col == "Name" and v.startswith("—"):
        return "rej", v  # explicit gap row: no senior contact in the CRM
    if v.startswith("http"):
        short = re.sub(r"^https?://(www\.)?linkedin\.com/in/", "", v).rstrip("/")
        return "", f'<a href="{v}">{html.escape(short)}</a>'
    return "", v


def render_sheet(ws, title: str) -> str:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""
    headers = [h for h in rows[0] if h]
    body: list[str] = []
    notes: list[str] = []

    for r in rows[1:]:
        vals = ["" if v is None else str(v) for v in r]
        if not any(vals):
            continue
        # Footnote rows sit in column A with the rest blank.
        if not any(vals[1:]):
            notes.append(vals[0])
            continue
        tds = []
        for i, col in enumerate(headers):
            klass, inner = cell(col, vals[i] if i < len(vals) else "")
            tds.append(f'<td class="{klass}">{inner}</td>' if klass else f"<td>{inner}</td>")
        body.append(f"<tr>{''.join(tds)}</tr>")

    th = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
    sub = " ".join(html.escape(n) for n in notes)
    return (
        f'<section><h2>{html.escape(title)}<span class="cnt">{len(body)}</span></h2>'
        + (f'<p class="sub">{sub}</p>' if sub else "")
        + f"<table><thead><tr>{th}</tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table></section>"
    )


def render(workbook_path: str, account: str) -> str:
    match = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(workbook_path))
    week = match.group(1) if match else "?"
    wb = load_workbook(workbook_path)

    parts = [
        render_sheet(wb[name], title) for name, title in SECTIONS if name in wb.sheetnames
    ]

    label = html.escape(account)
    out_path = os.path.join(
        os.path.dirname(workbook_path) or ".", f"weekly_view_{week}.html"
    )
    with open(out_path, "w") as fh:
        fh.write(
            "<!doctype html><html lang=en><head><meta charset=utf-8>"
            '<meta name=viewport content="width=device-width,initial-scale=1">'
            f"<title>Weekly Stakeholder Map — {label} — week of {week}</title>"
            f"<style>{CSS}</style></head><body>"
            f"<h1>Weekly Stakeholder Map — {label} · week of {week}</h1>"
            f'<div class="meta">Rendered from '
            f"{html.escape(os.path.basename(workbook_path))}. Contains contact data "
            "from the CRM — local file, not for publishing.</div>"
            f"{''.join(parts)}</body></html>"
        )
    return out_path
