"""Step 6 — emit. Four tabs, provenance on every row.

Nothing is written back to any system. The workbook and the JSON dump are the
only outputs, and the dump redacts MobilePhone to a boolean: sensitive contact
fields pass through to the CRM, never into a Quorom store.
"""

from __future__ import annotations

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..config import Config
from ..crm.fieldmap import NOT_AVAILABLE
from .stakeholders import NO_SENIOR_CONTACT

HEADER_FILL = "2F5B7C"


def _sheet(wb: Workbook, title: str, headers: list[str]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
    ws.freeze_panes = "A2"
    return ws


def _linkedin_cell(value) -> str:
    """None is not False. A CRM with no LinkedIn field says so in the cell, so a
    blank column cannot be read as 'nobody has one'."""
    if value is None:
        return NOT_AVAILABLE
    return "yes" if value else ""


def _crms_queried(cfg: Config) -> list[str]:
    """The CRMs this run actually called, in column order.

    Headers and the Source column are both built from this list, so a system
    that was never queried cannot appear in either. The names are the vendors'
    own, which is fine in a value that reports provenance — what is not fine is
    naming one that was not consulted.
    """
    names = []
    if cfg.hubspot.configured:
        names.append("hubspot")
    if cfg.salesforce.configured:
        names.append("salesforce")
    return names


def build_workbook(
    cfg: Config,
    reconciled: list[dict],
    coverage: list[dict],
    suppressed: list[str],
    stakeholders: list[dict],
    out_path: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Tab 1 — Met this week
    ws1 = _sheet(
        wb,
        "1 - Met this week",
        # "Title (CRM)", not "Title (SF)". The value already comes from either
        # CRM — Salesforce wins, HubSpot is the fallback — so "(SF)" was
        # imprecise even with Salesforce configured, and names a system that was
        # never called without it. Which system holds a differing title is
        # already stated in Flag, and the next column is "Mobile in CRM?".
        ["Name", "Email", "Title (CRM)", "LinkedIn?", "Mobile in CRM?", "Flag", "Source"],
    )
    for r in reconciled:
        ws1.append(
            [
                r.get("attendee_name"),
                r.get("email", ""),
                r.get("title", ""),
                _linkedin_cell(r.get("linkedin_in_crm")),
                "yes" if r.get("mobile_in_crm") else "GAP",
                r.get("flag", ""),
                "gong",
            ]
        )

    # Tab 2 — Missing from CRM
    #
    # A CRM that was not configured was not queried, so it gets no column at
    # all rather than a column of "not checked". Both are kept when both are
    # configured: which of the two holds the person is the point of this tab —
    # "in HubSpot but not Salesforce" is actionable, and a single merged
    # "In CRM?" would throw that away for anyone running both.
    hs_on = cfg.hubspot.configured
    sf_on = cfg.salesforce.configured
    source = "/".join(_crms_queried(cfg))
    ws2 = _sheet(
        wb,
        "2 - Missing from CRM",
        ["Name", "Email", "Company (domain)"]
        + (["In HubSpot?"] if hs_on else [])
        + (["In Salesforce?"] if sf_on else [])
        + ["Flag", "Source"],
    )
    for r in reconciled:
        in_sf = r.get("in_salesforce")
        in_hs = r.get("in_hubspot")
        # None is "not checked", and only ever arises for a CRM that is
        # unconfigured — whose column is not rendered. So a rendered cell is
        # always a real yes/no, and "not checked" never reaches this tab.
        if in_hs is False or in_sf is False:
            flag = r.get("flag", "")
            # "needs name/title" is redundant in a gap report — the row IS the gap.
            flag = flag if "shared inbox" in flag else ""
            row = [r.get("attendee_name"), r.get("email", ""), r.get("domain")]
            if hs_on:
                row.append("yes" if in_hs else "NO")
            if sf_on:
                row.append("yes" if in_sf else "NO")
            ws2.append(row + [flag, source])
    if suppressed:
        ws2.append([])
        ws2.append(
            [
                "Suppressed as non-contacts (no email/domain — likely meeting bots): "
                + ", ".join(suppressed)
            ]
        )

    # Tab 3 — Company coverage (triage)
    #
    # Same rule as tab 2, applied to counts: a provider that was not queried
    # contributes no column. A count column has to stay numeric to be sortable
    # and summable — writing "not checked" into it would turn the whole column
    # to text and quietly break sorting on the tab whose job is triage — so the
    # absence is expressed by dropping the column rather than by a value in it.
    ws3 = _sheet(
        wb,
        "3 - Company coverage",
        ["Company", "Company name", "Employees", "HQ", "Account type",
         "Meets profile?", "Met this wk"]
        + (["SF contacts", "SF focus-senior"] if sf_on else [])
        + (["HubSpot contacts"] if hs_on else []),
    )
    for c in sorted(coverage, key=lambda x: (not x.get("is_target"), -x.get("met", 0))):
        row = [
            c["domain"], c.get("name", ""), c.get("employees", ""), c.get("hq", ""),
            c.get("account_type", "") or "(blank)", c.get("meets", ""), c.get("met", 0),
        ]
        if sf_on:
            row += [c.get("sf_total", 0), c.get("sf_senior", 0)]
        if hs_on:
            row.append(c.get("hs_total", 0))
        ws3.append(row)
    ws3.append([])
    ws3.append(
        [
            "Meets profile? = the employee band and HQ geography from your focus "
            "profile. Account type is shown for context and is not used to filter."
        ]
    )

    # Tab 4 — Stakeholder list (the map)
    ws4 = _sheet(
        wb,
        "4 - Stakeholder list",
        ["Company", "Name", "Title", "Recent contact?", "LinkedIn", "Mobile in CRM?"],
    )
    for r in stakeholders:
        ws4.append(
            [r.get("company", ""), r.get("name", ""), r.get("title", ""),
             r.get("contact", ""), r.get("linkedin", ""), r.get("mobile", "")]
        )
    ws4.append([])
    # Two lines, and only what a reader needs to read a value in the table.
    # Design rationale, open questions and ticket references belong in the repo
    # and in Linear — not in a file that goes to a customer.
    ws4.append(
        [
            "People already in your CRM at these companies, most senior first. "
            "Others may exist who aren't in the CRM."
        ]
    )
    ws4.append(
        [
            f"Recent contact = a meeting, or activity logged in the CRM, in the last "
            f"{cfg.recent_days} days. Titles come from the CRM and may be out of date."
        ]
    )

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def dump_inputs(path: str, payload: dict) -> None:
    """Every input the run used, so the ordering can be re-tuned without going
    back to Salesforce."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w") as fh:
        json.dump(payload, fh, indent=2, default=str)


def count_gaps(stakeholders: list[dict]) -> int:
    return sum(1 for r in stakeholders if r.get("name") == NO_SENIOR_CONTACT)
