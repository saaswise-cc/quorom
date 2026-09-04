"""End-to-end over the two legs a session can actually reach: the product
database and the artifact. Salesforce and HubSpot are deliberately absent."""

from __future__ import annotations

import json

import psycopg
import pytest
import requests
from openpyxl import load_workbook

from quorom import db
from quorom.config import Config
from quorom.gong.importer import import_range
from quorom.weekly.run import run_weekly
from quorom.weekly.stakeholders import NO_SENIOR_CONTACT

ACCOUNT = "northwind.com"


def _seed_account(dsn: str) -> str:
    with psycopg.connect(dsn, autocommit=True) as conn:
        row = conn.execute(
            "insert into accounts (name, internal_domains) values (%s, %s) returning id",
            (ACCOUNT, ["northwind.com"]),
        ).fetchone()
    return str(row[0])


def _cfg(dsn: str, tmp_path, **overrides) -> Config:
    kwargs = dict(
        database_url=dsn,
        account=ACCOUNT,
        week_start="2026-08-17",
        output_dir=str(tmp_path),
    )
    kwargs.update(overrides)
    return Config(**kwargs)


def _seed_profile(dsn: str, account_id: str) -> None:
    """An active focus profile — run_weekly refuses to start without one."""
    with psycopg.connect(dsn, autocommit=True) as conn:
        conn.execute(
            "insert into user_focus_profiles (account_id, version_number, is_active, "
            "profile_data) values (%s, 1, true, %s)",
            (account_id, json.dumps({
                "employee_count_min": 200, "employee_count_max": 10000,
                "hq_geographies": ["North America"],
                "focus_seniority": ["c-level", "vp", "director"],
            })),
        )


def _import(dsn: str, account_id: str, gong_calls, database_conn=None):
    from tests.conftest import FakeGong

    with psycopg.connect(dsn) as conn:
        result = import_range(
            conn, FakeGong(gong_calls), account_id, ["northwind.com"],
            "2025-10-01", "2026-08-24", log=lambda *_: None,
        )
        conn.commit()
    return result


def test_import_classifies_and_resolves(database, gong_calls):
    account_id = _seed_account(database)
    result = _import(database, account_id, gong_calls)

    assert result.meetings_upserted == 3

    with psycopg.connect(database) as conn:
        kinds = dict(
            conn.execute(
                "select coalesce(domain_kind, '<null>'), count(*) from attendees "
                "group by 1"
            ).fetchall()
        )
        # Gong's affiliation wins where present; domain classification fills the
        # rest. The party with neither name nor email is not a person.
        assert kinds["internal"] == 2
        assert kinds["personal"] == 1          # sam@gmail.com
        assert kinds["<null>"] == 1            # named, no email, no affiliation
        assert kinds["external"] == 12         # Dana x2, support@, 9 trainees

        # Internal attendees never become people.
        emails = {
            r[0]
            for r in conn.execute(
                "select email from people where email is not null"
            ).fetchall()
        }
        assert "sam@northwind.com" not in emails
        assert {"dana.reyes@acme.com", "d.reyes@acme.com"} <= emails

        # The name-only attendee gets an unmatched person of its own.
        assert conn.execute(
            "select count(*) from people where unmatched"
        ).fetchone()[0] == 1


def test_import_times_the_run_it_just_did(database, gong_calls):
    """Counts alone cannot answer 'what would a year cost?'. The elapsed time
    is measured around the real work, not assembled by the caller."""
    account_id = _seed_account(database)
    result = _import(database, account_id, gong_calls)

    assert result.elapsed_seconds > 0
    assert str(result).endswith("elapsed")


def test_reimport_is_idempotent(database, gong_calls):
    """The failure this guards: an importer that inserts person_attendees with no uniqueness
    constraint. 0002 adds one, so a second run over an overlapping range would
    raise instead of doing nothing — a backfill that looks fine, then breaks the
    overnight job the next day."""
    account_id = _seed_account(database)
    first = _import(database, account_id, gong_calls)

    def counts() -> tuple:
        with psycopg.connect(database) as conn:
            return tuple(
                conn.execute(f"select count(*) from {t}").fetchone()[0]
                for t in ("meetings", "attendees", "people", "person_attendees",
                          "person_identifiers")
            )

    before = counts()
    second = _import(database, account_id, gong_calls)
    assert counts() == before
    assert second.calls_already_imported == 3
    assert second.attendees_created == 0
    assert first.meetings_upserted == second.meetings_upserted


def test_weekly_runs_without_a_crm(database, gong_calls, tmp_path):
    """No Salesforce, no HubSpot. The run must still produce the artifact, with
    the missing legs stated rather than silently absent."""
    account_id = _seed_account(database)
    _import(database, account_id, gong_calls)

    _seed_profile(database, account_id)

    cfg = _cfg(database, tmp_path)
    paths = run_weekly(cfg, log=lambda *_: None)

    wb = load_workbook(paths["xlsx"])
    assert wb.sheetnames == [
        "1 - Met this week",
        "2 - Missing from CRM",
        "3 - Company coverage",
        "4 - Stakeholder list",
    ]

    # 11 distinct external people met in the pinned week: Dana, support@, and
    # nine trainees. The historical call and the internal attendees are out.
    met = list(wb["1 - Met this week"].iter_rows(min_row=2, values_only=True))
    assert len(met) == 11

    # No CRM was queried, so nothing can be reported missing from one. The
    # columns that would have said so are not rendered at all — see
    # test_tabs_2_and_3_omit_a_crm_that_was_never_queried.
    missing = [
        r for r in wb["2 - Missing from CRM"].iter_rows(min_row=2, values_only=True)
        if r[0]
    ]
    assert missing == []

    dump = json.loads(open(paths["json"]).read())
    assert dump["account"] == ACCOUNT
    assert dump["focus_profile"]["employee_count_min"] == 200

    html = open(paths["html"]).read()
    assert "northwind.com" in html     # the account, not a hardcoded name
    assert "not for publishing" in html


def _headers(ws) -> list:
    return [h for h in next(ws.iter_rows(max_row=1, values_only=True)) if h]


def test_tabs_2_and_3_omit_a_crm_that_was_never_queried(database, gong_calls, tmp_path):
    """Four output surfaces, in the state the suite actually runs in.

    Nothing covered this before: the suite has always run with both CRMs off
    and only asserted tab 1, so a heading naming two vendors, two always-present
    columns, a hardcoded "hubspot/sfdc" provenance and a bare 0 contact count
    all stayed invisible. An unqueried provider must contribute no column, no
    count and no name anywhere a reader can see.
    """
    account_id = _seed_account(database)
    _import(database, account_id, gong_calls)
    _seed_profile(database, account_id)

    paths = run_weekly(_cfg(database, tmp_path), log=lambda *_: None)
    wb = load_workbook(paths["xlsx"])

    # Tab 2 — neither "In HubSpot?" nor "In Salesforce?" is offered, so
    # "not checked" never has to appear on this tab at all.
    assert _headers(wb["2 - Missing from CRM"]) == [
        "Name", "Email", "Company (domain)", "Flag", "Source",
    ]

    # Tab 3 — the count columns for both unqueried providers are gone. A 0
    # here would be indistinguishable from a company with genuinely no
    # contacts on file.
    assert _headers(wb["3 - Company coverage"]) == [
        "Company", "Company name", "Employees", "HQ", "Account type",
        "Meets profile?", "Met this wk",
    ]

    # No cell on any tab reports a check that never happened.
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(values_only=True):
            for value in row:
                assert "not checked" != str(value).strip().lower()

    # The inputs dump carries the same distinction as null, not zero, so the
    # JSON cannot be re-read later as "we looked and found none".
    dump = json.loads(open(paths["json"]).read())
    for company in dump["coverage"]:
        assert company["hs_total"] is None
        assert company["sf_total"] is None
        assert company["sf_senior"] is None

    # The HTML view renamed the tab to a heading naming both vendors. It is
    # the file a reader actually opens.
    html = open(paths["html"]).read()
    assert "Not in CRM" in html
    assert "Not in HubSpot or Salesforce" not in html
    assert "hubspot" not in html.lower()


@pytest.mark.parametrize(
    "hs_on, sf_on, crm_columns, source",
    [
        (True, True, ["In HubSpot?", "In Salesforce?"], "hubspot/salesforce"),
        (True, False, ["In HubSpot?"], "hubspot"),
        (False, True, ["In Salesforce?"], "salesforce"),
        (False, False, [], ""),
    ],
)
def test_workbook_columns_follow_the_crms_configured(
    tmp_path, hs_on, sf_on, crm_columns, source
):
    """Both columns survive when both CRMs are on — "in HubSpot but not
    Salesforce" is the actionable answer and merging them into one "In CRM?"
    would lose it. Only the unconfigured one is dropped, and Source names
    exactly what was queried.

    A unit test rather than an end-to-end one: configuring a CRM leg here would
    make the run reach for the real API. The credentials below are literals, so
    nothing real can leak into a failure message.
    """
    from quorom.config import HubSpotConfig, SalesforceConfig
    from quorom.weekly.workbook import build_workbook

    cfg = Config(
        account=ACCOUNT,
        output_dir=str(tmp_path),
        hubspot=HubSpotConfig(api_key="test-key" if hs_on else ""),
        salesforce=SalesforceConfig(
            access_token="test-token" if sf_on else "",
            instance_url="https://example.invalid" if sf_on else "",
        ),
    )
    assert cfg.hubspot.configured is hs_on and cfg.salesforce.configured is sf_on

    reconciled = [
        {
            "attendee_name": "Dana Reyes", "email": "dana@acme.com",
            "domain": "acme.com", "flag": "", "title": "Director of RevOps",
            "mobile_in_crm": False, "linkedin_in_crm": False,
            # False is a real answer from a CRM that was called; None is what
            # reconcile() writes for one that was not.
            "in_hubspot": False if hs_on else None,
            "in_salesforce": False if sf_on else None,
        }
    ]
    coverage = [
        {
            "domain": "acme.com", "name": "Acme", "employees": 500, "hq": "US",
            "account_type": "", "meets": "yes", "met": 1, "is_target": True,
            "sf_total": 4 if sf_on else None,
            "sf_senior": 1 if sf_on else None,
            "hs_total": 7 if hs_on else None,
        }
    ]

    out = str(tmp_path / "wb.xlsx")
    build_workbook(cfg, reconciled, coverage, [], [], out)
    wb = load_workbook(out)

    ws2 = wb["2 - Missing from CRM"]
    assert _headers(ws2) == (
        ["Name", "Email", "Company (domain)"] + crm_columns + ["Flag", "Source"]
    )

    ws3 = wb["3 - Company coverage"]
    assert _headers(ws3) == (
        ["Company", "Company name", "Employees", "HQ", "Account type",
         "Meets profile?", "Met this wk"]
        + (["SF contacts", "SF focus-senior"] if sf_on else [])
        + (["HubSpot contacts"] if hs_on else [])
    )

    rows = [r for r in ws2.iter_rows(min_row=2, values_only=True) if r[0]]
    if not crm_columns:
        # Nothing can be missing from a CRM that was never consulted.
        assert rows == []
        return

    assert len(rows) == 1
    # Source names every CRM queried and nothing else.
    assert rows[0][-1] == source
    for vendor in ("hubspot", "salesforce"):
        if vendor not in source:
            assert vendor not in " ".join(str(v) for v in rows[0]).lower()
    # Every rendered CRM cell is a real answer, never "not checked".
    assert list(rows[0][3:3 + len(crm_columns)]) == ["NO"] * len(crm_columns)


def test_the_map_filter_keeps_a_company_it_could_not_assess():
    """The filter half. `is_target` is False for a company that failed the ICP
    test AND for one the test could not run on, so filtering on it alone drops
    the second kind off tab 4 — a company disappearing from the map because
    data nobody fetched did not clear a bar."""
    from quorom.weekly.stakeholders import companies_for_map

    coverage = [
        {"domain": "target.com", "assessed": True, "is_target": True},
        {"domain": "rejected.com", "assessed": True, "is_target": False},
        {"domain": "unassessed.com", "assessed": False, "is_target": False},
    ]

    kept = [c["domain"] for c in companies_for_map(coverage)]

    assert kept == ["target.com", "unassessed.com"]
    # A company the test rejected is a decision, and stays off the map.
    assert "rejected.com" not in kept


def test_no_crm_does_not_silently_empty_the_stakeholder_map(
    database, gong_calls, tmp_path
):
    """With no CRM, the ICP test cannot run — and must say so rather than
    reporting a verdict.

    Before this, empty firmographics made every company fail on "no size", and
    the same verdict is the filter feeding tab 4, so the map came out empty.
    Nothing errored and the workbook had its usual shape: an empty tab 4 reads
    as "nobody worth considering this week", which is a finding a reader would
    act on, rather than "the test never ran".
    """
    from quorom.weekly.coverage import NOT_ASSESSED
    from quorom.weekly.stakeholders import ICP_NOT_ASSESSED

    account_id = _seed_account(database)
    _import(database, account_id, gong_calls)
    _seed_profile(database, account_id)

    paths = run_weekly(_cfg(database, tmp_path), log=lambda *_: None)
    wb = load_workbook(paths["xlsx"])

    # Tab 3 — the verdict column states that no verdict was reached. Not "no
    # size", which is a finding about the company.
    ws3 = wb["3 - Company coverage"]
    verdict = _headers(ws3).index("Meets profile?")
    companies = [r for r in ws3.iter_rows(min_row=2, values_only=True) if r[1] or r[6]]
    assert companies, "the company met this week must still appear on tab 3"
    for row in companies:
        assert row[verdict] == NOT_ASSESSED

    # Tab 4 — not empty. Every company that reached the map is on it, saying
    # why there are no people rather than being absent.
    tab4 = [
        r for r in wb["4 - Stakeholder list"].iter_rows(min_row=2, values_only=True)
        if r[1]
    ]
    assert tab4, "tab 4 must not be empty when the ICP test could not run"
    assert {r[1] for r in tab4} == {ICP_NOT_ASSESSED}
    assert {r[0] for r in tab4} == {"acme.com"}
    # Not the "we looked and found nobody" row — nothing was looked at.
    assert all(r[1] != NO_SENIOR_CONTACT for r in tab4)

    # Tab 1 — the header no longer names a CRM this run never called.
    assert "Title (CRM)" in _headers(wb["1 - Met this week"])
    assert "Title (SF)" not in _headers(wb["1 - Met this week"])

    # The dump carries the third state as its own field, so "could not assess"
    # cannot later be re-read as "assessed and rejected".
    dump = json.loads(open(paths["json"]).read())
    for company in dump["coverage"]:
        assert company["assessed"] is False
        assert company["is_target"] is False
        assert company["meets"] == NOT_ASSESSED
    # Nothing was queried, so no bench is claimed for it — not even an empty one.
    assert dump["sf_bench"] == []

    # The HTML must not colour a test that never ran as a rejection.
    html = open(paths["html"]).read()
    assert "not assessed" in html
    assert f'class="na">{NOT_ASSESSED}' in html
    assert f'class="rej">{NOT_ASSESSED}' not in html


def test_history_splits_on_a_changed_address(database, gong_calls):
    """The defect that justifies the identity tables, as a test rather than a
    claim: Dana attended under two addresses, so the email-keyed history query
    reports her as two people with two different last_met dates."""
    account_id = _seed_account(database)
    _import(database, account_id, gong_calls)

    cfg = Config(database_url=database, account=ACCOUNT, week_start="2026-08-17")
    with psycopg.connect(database) as conn:
        history = db.met_history(conn, cfg, ["acme.com"])

        assert history["dana.reyes@acme.com"]["last_met"].isoformat() == "2026-08-18"
        assert history["d.reyes@acme.com"]["last_met"].isoformat() == "2025-10-21"

        # person_identifiers already holds what would collapse them.
        rows = conn.execute(
            "select count(distinct person_id) from person_identifiers "
            "where email in ('dana.reyes@acme.com', 'd.reyes@acme.com')"
        ).fetchone()[0]
    # Two records today because the addresses were never linked; the read path
    # that fixes this is the one described in migrations/0002_identity.sql.
    assert rows == 2


def test_group_call_is_labelled_not_judged(database, gong_calls):
    account_id = _seed_account(database)
    _import(database, account_id, gong_calls)

    cfg = Config(database_url=database, account=ACCOUNT, week_start="2026-08-17")
    with psycopg.connect(database) as conn:
        history = db.met_history(conn, cfg, ["acme.com"])

    from quorom.weekly.stakeholders import recent_contact

    trainee = history["trainee1@acme.com"]
    assert trainee["smallest_meeting"] == 9      # above GROUP_CALL_MIN of 8
    # Recency is computed against today, so assert the label rather than the date.
    label = recent_contact(cfg, trainee, None)
    assert "group call, 9 attendees" in label or label.startswith("no —")


@pytest.mark.parametrize(
    "title, rank",
    [
        ("Chief Revenue Officer", 3),
        ("Vice President, Sales", 2),   # must not match on 'president'
        ("SVP Marketing", 2),
        ("Chief of Staff", 2),          # not C-level here
        ("Director of RevOps", 1),
        ("", 0),
    ],
)
def test_seniority_ordering(title, rank):
    from quorom.weekly.stakeholders import seniority_rank

    assert seniority_rank(title) == rank


def test_soql_quoting():
    from quorom.crm.salesforce import soql_quote

    assert soql_quote("o'brien@acme.com") == "o\\'brien@acme.com"
    assert soql_quote("a\\b") == "a\\\\b"


# --- HubSpot 429 handling ---------------------------------------------------- #
#
# The search endpoint is burst-limited, so a weekly run has to wait one out
# rather than die. No network: requests.post is replaced with a scripted
# sequence of responses, and the sleep is recorded instead of taken.


class _Resp:
    def __init__(self, status, body=None, headers=None):
        self.status_code = status
        self._body = body or {}
        self.headers = headers or {}

    def json(self):
        return self._body

    def raise_for_status(self):
        if self.status_code >= 400:
            raise requests.HTTPError(f"{self.status_code}", response=self)


@pytest.fixture
def hubspot_calls(monkeypatch):
    """Script HubSpot's replies; return the recorded sleeps and request count."""
    from quorom.crm import hubspot as hs_mod

    state = {"sent": [], "slept": []}

    def install(responses):
        queue = list(responses)

        def fake_post(url, **kwargs):
            state["sent"].append(kwargs.get("json"))
            return queue.pop(0)

        monkeypatch.setattr(hs_mod.requests, "post", fake_post)
        monkeypatch.setattr(hs_mod.time, "sleep", lambda s: state["slept"].append(s))
        return state

    return install


def _client():
    """A client with a fake key, built without touching Config().

    Config() calls load_dotenv(), so constructing one here would pull the real
    HubSpot key out of .env and put it in any failure output.
    """
    from types import SimpleNamespace

    from quorom.config import HubSpotConfig
    from quorom.crm.hubspot import HubSpot

    return HubSpot(SimpleNamespace(hubspot=HubSpotConfig(api_key="test-key")))


def test_hubspot_retries_a_429_then_succeeds(hubspot_calls):
    hit = {"results": [{"id": "1", "properties": {"email": "a@acme.com"}}]}
    state = hubspot_calls([_Resp(429), _Resp(200, hit)])

    got = _client().contact_by_email("a@acme.com")

    assert got.email == "a@acme.com"             # the run continues
    assert len(state["sent"]) == 2               # it retried exactly once
    assert state["slept"] == [2.0]               # backed off before retrying


def test_hubspot_honours_retry_after(hubspot_calls):
    state = hubspot_calls(
        [_Resp(429, headers={"Retry-After": "7"}), _Resp(200, {"total": 4})]
    )

    assert _client().count_domain("acme.com") == 4
    assert state["slept"] == [7.0]               # HubSpot's number, not ours


def test_hubspot_caps_a_huge_retry_after(hubspot_calls):
    from quorom.crm.hubspot import MAX_BACKOFF

    state = hubspot_calls(
        [_Resp(429, headers={"Retry-After": "86400"}), _Resp(200, {"total": 0})]
    )

    _client().count_domain("acme.com")
    assert state["slept"] == [MAX_BACKOFF]        # bounded, so it cannot hang


def test_hubspot_gives_up_loudly_when_throttling_persists(hubspot_calls):
    from quorom.crm.hubspot import HubSpotRateLimited, MAX_ATTEMPTS

    state = hubspot_calls([_Resp(429)] * MAX_ATTEMPTS)

    with pytest.raises(HubSpotRateLimited):
        _client().contact_by_email("a@acme.com")

    assert len(state["sent"]) == MAX_ATTEMPTS     # capped, not unbounded
    assert len(state["slept"]) == MAX_ATTEMPTS - 1  # no sleep after the last try


def test_hubspot_does_not_retry_other_errors(hubspot_calls):
    state = hubspot_calls([_Resp(500)])

    with pytest.raises(requests.HTTPError):
        _client().contact_by_email("a@acme.com")

    assert len(state["sent"]) == 1                # 500 still fails immediately
